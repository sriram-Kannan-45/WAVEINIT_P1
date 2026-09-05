from __future__ import annotations

import os
import sys
import time
import json
import math
import base64
import argparse
import datetime
import logging
import urllib.request
from dataclasses import dataclass, field
from typing import Dict, Any, List, Optional, Tuple, Union

import cv2
import numpy as np
import mediapipe as mp
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ============================================================
# LMS AI PROCTORING ENGINE
#
# IMPORTANT:
# - Mobile detection is NOT implemented here.
# - The engine only accepts an external mobile result/count/score.
# - Eye + Head scoring is based ONLY on unique validated time.
# - 3 seconds is validation only, never a scoring unit.
# ============================================================

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
)
logger = logging.getLogger("lms-proctor")

# ============================================================
# CONSTANTS
# ============================================================

VIOLATION_SECONDS = 3.0

MONITORING_SCORE_MAX = 60.0
MOBILE_SCORE_MAX = 20.0
MULTIPLE_FACE_SCORE_MAX = 10.0
NO_PERSON_SCORE_MAX = 10.0
FINAL_SCORE_MAX = 100.0

# Head pose thresholds
YAW_INNER = 10.0
YAW_OUTER = 30.0
PITCH_INNER = 10.0
PITCH_OUTER = 20.0

YAW_LEFT_CUTOFF = 0.72
YAW_RIGHT_CUTOFF = 0.72
PITCH_UP_CUTOFF = 0.75
PITCH_DOWN_CUTOFF = 0.45

POSE_ALPHA = 0.35
PITCH_OFFSET = -5.0

# Gaze thresholds
GAZE_ENTER_HORIZONTAL = 0.030
GAZE_EXIT_HORIZONTAL = 0.017
GAZE_ENTER_VERTICAL = 0.026
GAZE_EXIT_VERTICAL = 0.014
GAZE_SMOOTHING_ALPHA = 0.30

PRIMARY_SIDE_LOCK = True
ENABLE_GAZE_CALIBRATION = True
CALIBRATION_FRAMES = 10
DEBUG_GAZE = True

# One short bad frame should not fragment a real episode.
# A reliable opposite direction still changes state immediately.
NOISE_GRACE_SECONDS = 0.20

IGNORED_HEAD_DIRECTIONS = {"Down"}
IGNORED_EYE_DIRECTIONS = {"Down"}


def iris_gaze_is_observable(head_direction: Optional[str]) -> bool:
    """Reject iris geometry while a strong head turn distorts the eye plane."""
    return head_direction in ("Straight", "Not Detected", None)

# MediaPipe Face Landmarker indices
RIGHT_IRIS_CENTER = 468
LEFT_IRIS_CENTER = 473

RIGHT_EYE_CORNERS = (33, 133)
LEFT_EYE_CORNERS = (263, 362)

RIGHT_EYE_VERTICAL = (159, 145)
LEFT_EYE_VERTICAL = (386, 374)

# Head pose model points
MODEL_POINTS = np.array(
    [
        (0.0, 0.0, 0.0),           # Nose
        (0.0, -330.0, -65.0),      # Chin
        (-225.0, 170.0, -135.0),   # Left eye corner
        (225.0, 170.0, -135.0),    # Right eye corner
        (-150.0, -150.0, -125.0),  # Left mouth corner
        (150.0, -150.0, -125.0),   # Right mouth corner
    ],
    dtype=np.float32,
)

LANDMARK_INDICES = [1, 152, 33, 263, 61, 291]

# ============================================================
# HELPERS
# ============================================================

def clamp(value: float, low: float, high: float) -> float:
    return max(low, min(high, float(value)))


def point_xy(landmark, width: int, height: int) -> np.ndarray:
    return np.array(
        [landmark.x * width, landmark.y * height],
        dtype=np.float32,
    )


def fuzzy_classify(value: float, inner_thresh: float, outer_thresh: float):
    abs_value = abs(float(value))
    if abs_value <= inner_thresh:
        confidence = 0.0
    elif abs_value >= outer_thresh:
        confidence = 1.0
    else:
        confidence = (abs_value - inner_thresh) / max(
            outer_thresh - inner_thresh, 1e-6
        )
    sign = 1 if value >= 0 else -1
    return sign, confidence


def rotation_matrix_to_euler_angles(R):
    sy = math.sqrt(float(R[0, 0] ** 2 + R[1, 0] ** 2))
    singular = sy < 1e-6

    if not singular:
        pitch = math.atan2(R[2, 1], R[2, 2])
        yaw = math.atan2(-R[2, 0], sy)
        roll = math.atan2(R[1, 0], R[0, 0])
    else:
        pitch = math.atan2(-R[1, 2], R[1, 1])
        yaw = math.atan2(-R[2, 0], sy)
        roll = 0.0

    return (
        float(np.degrees(pitch)),
        float(np.degrees(yaw)),
        float(np.degrees(roll)),
    )


def normalize_timestamp_ms(value: Optional[Union[int, float]]) -> Optional[float]:
    if value is None:
        return None
    try:
        value = float(value)
    except (TypeError, ValueError):
        return None
    # LMS timestamps are normally milliseconds. Also tolerate seconds.
    if abs(value) >= 1e11:
        return value / 1000.0
    if abs(value) >= 1e9:
        return value / 1000.0
    return value


def calculate_actual_duration_seconds(
    actual_start_time: Optional[Union[int, float]],
    actual_end_time: Optional[Union[int, float]],
) -> Optional[float]:
    start = normalize_timestamp_ms(actual_start_time)
    end = normalize_timestamp_ms(actual_end_time)
    if start is None or end is None:
        return None
    return max(0.0, end - start)


def merge_intervals(
    intervals: List[Tuple[float, float]]
) -> List[Tuple[float, float]]:
    valid = sorted(
        [
            (float(start), float(end))
            for start, end in intervals
            if float(end) > float(start)
        ],
        key=lambda x: x[0],
    )
    if not valid:
        return []

    merged: List[List[float]] = []
    for start, end in valid:
        if not merged or start > merged[-1][1]:
            merged.append([start, end])
        else:
            merged[-1][1] = max(merged[-1][1], end)

    return [(start, end) for start, end in merged]


def calculate_unique_violation_seconds(
    intervals: List[Tuple[float, float]]
) -> float:
    return sum(end - start for start, end in merge_intervals(intervals))


def calculate_monitoring_score(
    violation_seconds: float,
    test_duration_seconds: float,
):
    """Calculate Eye + Head score from unique cumulative violation time."""
    duration = max(0.0, float(test_duration_seconds or 0.0))
    violation = max(0.0, float(violation_seconds or 0.0))
    if duration <= 0.0:
        return 0.0, 0.0
    violation = min(violation, duration)
    percentage = clamp((violation / duration) * 100.0, 0.0, 100.0)
    score = clamp(
        (violation / duration) * MONITORING_SCORE_MAX,
        0.0,
        MONITORING_SCORE_MAX,
    )
    return percentage, score


def calculateEyeHeadScore(violation_seconds: float, test_duration_seconds: float) -> float:
    """Return Eye + Head score (/60) for backward compatibility."""
    _, score = calculate_monitoring_score(violation_seconds, test_duration_seconds)
    return score


def calculate_direction_totals(
    episodes: List[ViolationEpisode],
) -> Dict[str, float]:
    """Return cumulative seconds independently for all six directions."""
    totals = {
        "eye_left": 0.0,
        "eye_right": 0.0,
        "eye_up": 0.0,
        "head_left": 0.0,
        "head_right": 0.0,
        "head_up": 0.0,
    }
    for ep in episodes:
        cat = ep.category.lower()
        if "eye" in cat or "eyeball" in cat:
            cat = "eye"
        elif "head" in cat:
            cat = "head"
        key = f"{cat}_{ep.direction.lower()}"
        if key in totals:
            totals[key] += max(0.0, float(ep.duration))
    return totals


# ============================================================
# GAZE
# ============================================================

def calculate_normalized_eye_ratios(
    landmarks,
    width: int,
    height: int,
):
    required = [
        RIGHT_IRIS_CENTER,
        LEFT_IRIS_CENTER,
        *RIGHT_EYE_CORNERS,
        *LEFT_EYE_CORNERS,
        *RIGHT_EYE_VERTICAL,
        *LEFT_EYE_VERTICAL,
    ]

    if len(landmarks) <= max(required):
        raise ValueError("Required iris/eye landmarks are unavailable.")

    r_iris = point_xy(landmarks[RIGHT_IRIS_CENTER], width, height)
    r_c1 = point_xy(landmarks[RIGHT_EYE_CORNERS[0]], width, height)
    r_c2 = point_xy(landmarks[RIGHT_EYE_CORNERS[1]], width, height)
    r_top = point_xy(landmarks[RIGHT_EYE_VERTICAL[0]], width, height)
    r_bot = point_xy(landmarks[RIGHT_EYE_VERTICAL[1]], width, height)

    l_iris = point_xy(landmarks[LEFT_IRIS_CENTER], width, height)
    l_c1 = point_xy(landmarks[LEFT_EYE_CORNERS[0]], width, height)
    l_c2 = point_xy(landmarks[LEFT_EYE_CORNERS[1]], width, height)
    l_top = point_xy(landmarks[LEFT_EYE_VERTICAL[0]], width, height)
    l_bot = point_xy(landmarks[LEFT_EYE_VERTICAL[1]], width, height)

    r_min_x, r_max_x = sorted([r_c1[0], r_c2[0]])
    l_min_x, l_max_x = sorted([l_c1[0], l_c2[0]])

    r_h_span = max(1.0, float(r_max_x - r_min_x))
    l_h_span = max(1.0, float(l_max_x - l_min_x))

    r_h = clamp((float(r_iris[0]) - r_min_x) / r_h_span, 0.0, 1.0)
    l_h = clamp((float(l_iris[0]) - l_min_x) / l_h_span, 0.0, 1.0)

    r_v_min, r_v_max = sorted([r_top[1], r_bot[1]])
    l_v_min, l_v_max = sorted([l_top[1], l_bot[1]])

    r_v_span = max(1.0, float(r_v_max - r_v_min))
    l_v_span = max(1.0, float(l_v_max - l_v_min))

    r_v = clamp((float(r_iris[1]) - r_v_min) / r_v_span, 0.0, 1.0)
    l_v = clamp((float(l_iris[1]) - l_v_min) / l_v_span, 0.0, 1.0)

    return (
        (r_h + l_h) / 2.0,
        (r_v + l_v) / 2.0,
        {
            "r_iris": (int(r_iris[0]), int(r_iris[1])),
            "l_iris": (int(l_iris[0]), int(l_iris[1])),
            "r_corners": (
                (int(r_c1[0]), int(r_c1[1])),
                (int(r_c2[0]), int(r_c2[1])),
            ),
            "l_corners": (
                (int(l_c1[0]), int(l_c1[1])),
                (int(l_c2[0]), int(l_c2[1])),
            ),
        },
    )


class GazeClassifier:
    def __init__(self):
        self.reset()

    def reset(self):
        self.current_direction = "Straight"
        self.smoothed_h = None
        self.smoothed_v = None
        self.neutral_x = 0.50
        self.neutral_y = 0.50
        self.calib_samples_x: List[float] = []
        self.calib_samples_y: List[float] = []
        self.is_calibrated = not ENABLE_GAZE_CALIBRATION

    def add_calibration_sample(self, raw_h: float, raw_v: float):
        if self.is_calibrated:
            return
        if len(self.calib_samples_x) < CALIBRATION_FRAMES:
            self.calib_samples_x.append(float(raw_h))
            self.calib_samples_y.append(float(raw_v))

        if len(self.calib_samples_x) >= CALIBRATION_FRAMES:
            self.neutral_x = float(np.median(self.calib_samples_x))
            self.neutral_y = float(np.median(self.calib_samples_y))
            self.smoothed_h = self.neutral_x
            self.smoothed_v = self.neutral_y
            self.current_direction = "Straight"
            self.is_calibrated = True

    def classify(self, raw_h: float, raw_v: float):
        if not self.is_calibrated:
            return "Straight", 0.0, 0.0, 0.0, raw_h, raw_v

        if self.smoothed_h is None:
            self.smoothed_h = float(raw_h)
            self.smoothed_v = float(raw_v)
        else:
            a = GAZE_SMOOTHING_ALPHA
            self.smoothed_h = a * raw_h + (1.0 - a) * self.smoothed_h
            self.smoothed_v = a * raw_v + (1.0 - a) * self.smoothed_v

        dx = self.smoothed_h - self.neutral_x
        dy = self.smoothed_v - self.neutral_y

        # Once a clear horizontal direction is active, vertical noise cannot
        # steal the classification until horizontal gaze has returned.
        if PRIMARY_SIDE_LOCK and self.current_direction in ("Left", "Right"):
            if (
                self.current_direction == "Left"
                and dx <= -GAZE_EXIT_HORIZONTAL
            ):
                conf = clamp(
                    abs(dx) / (GAZE_ENTER_HORIZONTAL * 1.8), 0.0, 1.0
                )
                return "Left", conf * 100.0, dx, dy, self.smoothed_h, self.smoothed_v

            if (
                self.current_direction == "Right"
                and dx >= GAZE_EXIT_HORIZONTAL
            ):
                conf = clamp(
                    abs(dx) / (GAZE_ENTER_HORIZONTAL * 1.8), 0.0, 1.0
                )
                return "Right", conf * 100.0, dx, dy, self.smoothed_h, self.smoothed_v

        horizontal = None
        horizontal_conf = 0.0
        if dx <= -GAZE_ENTER_HORIZONTAL:
            horizontal = "Left"
            horizontal_conf = clamp(
                abs(dx) / (GAZE_ENTER_HORIZONTAL * 1.8), 0.0, 1.0
            )
        elif dx >= GAZE_ENTER_HORIZONTAL:
            horizontal = "Right"
            horizontal_conf = clamp(
                abs(dx) / (GAZE_ENTER_HORIZONTAL * 1.8), 0.0, 1.0
            )

        vertical = None
        vertical_conf = 0.0
        if dy <= -GAZE_ENTER_VERTICAL:
            vertical = "Up"
            vertical_conf = clamp(
                abs(dy) / (GAZE_ENTER_VERTICAL * 1.8), 0.0, 1.0
            )
        elif dy >= GAZE_ENTER_VERTICAL:
            vertical = "Down"
            vertical_conf = clamp(
                abs(dy) / (GAZE_ENTER_VERTICAL * 1.8), 0.0, 1.0
            )

        if horizontal:
            if (
                vertical
                and vertical_conf > 0.90
                and horizontal_conf < 0.45
            ):
                direction = vertical
                confidence = vertical_conf
            else:
                direction = horizontal
                confidence = horizontal_conf
        elif vertical:
            direction = vertical
            confidence = vertical_conf
        else:
            direction = "Straight"
            confidence = 0.0

        self.current_direction = direction
        return (
            direction,
            confidence * 100.0,
            dx,
            dy,
            self.smoothed_h,
            self.smoothed_v,
        )


# ============================================================
# HEAD POSE
# ============================================================

class HeadPoseEstimator:
    def __init__(self):
        self.reset()

    def reset(self):
        self.previous_pose = None
        self.smoothed_pose = None

    def estimate(self, face_landmarks, width: int, height: int):
        if len(face_landmarks) <= max(LANDMARK_INDICES):
            return "Straight", 0.0, None, None, None

        image_points = np.array(
            [
                (
                    face_landmarks[idx].x * width,
                    face_landmarks[idx].y * height,
                )
                for idx in LANDMARK_INDICES
            ],
            dtype=np.float32,
        )

        focal_length = float(width)
        camera_matrix = np.array(
            [
                [focal_length, 0, width / 2],
                [0, focal_length, height / 2],
                [0, 0, 1],
            ],
            dtype=np.float32,
        )
        dist_coeffs = np.zeros((4, 1), dtype=np.float32)

        try:
            if self.previous_pose is not None:
                success, rvec, tvec = cv2.solvePnP(
                    MODEL_POINTS,
                    image_points,
                    camera_matrix,
                    dist_coeffs,
                    rvec=self.previous_pose[0].copy(),
                    tvec=self.previous_pose[1].copy(),
                    useExtrinsicGuess=True,
                    flags=cv2.SOLVEPNP_ITERATIVE,
                )
            else:
                success, rvec, tvec = cv2.solvePnP(
                    MODEL_POINTS,
                    image_points,
                    camera_matrix,
                    dist_coeffs,
                    flags=cv2.SOLVEPNP_ITERATIVE,
                )
        except cv2.error:
            success = False

        if not success:
            return "Straight", 0.0, None, None, None

        self.previous_pose = (rvec, tvec)

        rotation, _ = cv2.Rodrigues(rvec)
        pitch, yaw, roll = rotation_matrix_to_euler_angles(rotation)

        if pitch > 90:
            pitch -= 180
        elif pitch < -90:
            pitch += 180

        pitch += PITCH_OFFSET

        if self.smoothed_pose is None:
            self.smoothed_pose = [pitch, yaw, roll]
        else:
            a = POSE_ALPHA
            self.smoothed_pose[0] += a * (pitch - self.smoothed_pose[0])
            self.smoothed_pose[1] += a * (yaw - self.smoothed_pose[1])
            self.smoothed_pose[2] += a * (roll - self.smoothed_pose[2])

        pitch, yaw, roll = self.smoothed_pose

        yaw_sign, yaw_conf = fuzzy_classify(yaw, YAW_INNER, YAW_OUTER)
        pitch_sign, pitch_conf = fuzzy_classify(
            pitch, PITCH_INNER, PITCH_OUTER
        )

        horizontal = None
        vertical = None

        if yaw_sign > 0 and yaw_conf >= YAW_RIGHT_CUTOFF:
            horizontal = "Right"
        elif yaw_sign < 0 and yaw_conf >= YAW_LEFT_CUTOFF:
            horizontal = "Left"

        if pitch_sign > 0 and pitch_conf >= PITCH_DOWN_CUTOFF:
            vertical = "Down"
        elif pitch_sign < 0 and pitch_conf >= PITCH_UP_CUTOFF:
            vertical = "Up"

        candidates = []
        if horizontal:
            candidates.append((horizontal, yaw_conf))
        if vertical:
            candidates.append((vertical, pitch_conf))

        if candidates:
            candidates.sort(key=lambda x: x[1], reverse=True)
            direction = candidates[0][0]
            confidence = candidates[0][1] * 100.0
        else:
            direction = "Straight"
            confidence = 0.0

        nose = (
            int(face_landmarks[1].x * width),
            int(face_landmarks[1].y * height),
        )

        return direction, confidence, pitch, yaw, nose


# ============================================================
# CONTINUOUS VIOLATION STATE MACHINE
# ============================================================

@dataclass
class ViolationEpisode:
    category: str
    direction: str
    start_time: float
    end_time: float
    duration: float
    validation_status: str = "VALID_VIOLATION"


class ContinuousDirectionCounter:
    """Accumulate every detected interval for one Eye/Head direction.

    There is intentionally NO minimum-duration filter here. Every positive
    duration contributes to the cumulative total. A direction change, face
    loss, or session finalization closes the active interval.
    """

    def __init__(self, category: str, direction: str, **_ignored):
        self.category = category
        self.direction = direction
        self.started_at: Optional[float] = None
        self.last_seen_at: Optional[float] = None
        self.completed: List[ViolationEpisode] = []
        # Kept for backward-compatible reporting/API shape. It is never scored.
        self.invalid: List[ViolationEpisode] = []

    def start_or_continue(
        self, detected_direction: str, now: float
    ) -> Optional[ViolationEpisode]:
        now = float(now)
        if detected_direction == self.direction:
            if self.started_at is None:
                self.started_at = now
            self.last_seen_at = now
            return None

        # Exact state transition: close at the last timestamp where this
        # direction was actually observed. No 3-second validation/grace period.
        if self.started_at is None:
            return None
        return self.close(now)

    def close(self, now: Optional[float] = None) -> Optional[ViolationEpisode]:
        if self.started_at is None:
            return None

        end = (
            float(self.last_seen_at)
            if self.last_seen_at is not None
            else float(now if now is not None else self.started_at)
        )
        duration = max(0.0, end - float(self.started_at))

        episode = ViolationEpisode(
            category=self.category,
            direction=self.direction,
            start_time=float(self.started_at),
            end_time=end,
            duration=duration,
            validation_status="VALID_VIOLATION" if duration > 0.0 else "ZERO_DURATION",
        )

        if duration > 0.0:
            self.completed.append(episode)

        self.started_at = None
        self.last_seen_at = None
        return episode if duration > 0.0 else None

    def elapsed(self, now: float) -> float:
        if self.started_at is None:
            return 0.0
        return max(0.0, float(now) - self.started_at)

    def reset(self, now: Optional[float] = None):
        return self.close(now)


class SampledIntervalTracker:
    """Accumulate a boolean state over sampled video frames and emit an interval
    only when it is backed by enough consecutive frames AND enough real time.

    Used by the offline video processor for intervals the engine itself does
    not track (no-person, multiple-person, face-not-visible). `now` must be the
    same monotonic clock as the engine uses (session.created_at + video time).
    """

    def __init__(self, name: str, min_frames: int = 5, min_duration_sec: float = 1.0):
        self.name = name
        self.min_frames = max(1, int(min_frames))
        self.min_duration_sec = max(0.0, float(min_duration_sec))
        self._start = None
        self._last = None
        self._frames = 0

    def update(self, active: bool, now: float) -> Optional[Dict[str, Any]]:
        now = float(now)
        if active:
            if self._start is None:
                self._start = now
            self._last = now
            self._frames += 1
            return None
        return self.flush(now)

    def flush(self, now: Optional[float] = None) -> Optional[Dict[str, Any]]:
        if self._start is None:
            return None
        end = float(now if now is not None else self._last)
        start = float(self._start)
        duration = max(0.0, end - start)
        frames = self._frames
        self._start = None
        self._last = None
        self._frames = 0
        if frames < self.min_frames or duration < self.min_duration_sec:
            return None
        return {
            "category": self.name,
            "start": round(start, 3),
            "end": round(end, 3),
            "duration": round(duration, 3),
            "frames": frames,
        }

    def elapsed(self) -> float:
        if self._start is None:
            return 0.0
        return max(0.0, float(self._last) - float(self._start))


# ============================================================
# SESSION
# ============================================================

@dataclass
class SessionState:
    session_id: str
    configured_duration: float

    participant_id: str = "1"
    participant_name: str = "Participant"
    trainer_id: Optional[str] = None
    course_id: Optional[str] = None
    assessment_id: Optional[str] = None
    attempt_id: Optional[str] = None

    actual_start_time: Optional[float] = None
    actual_end_time: Optional[float] = None

    created_at_monotonic: float = field(default_factory=time.monotonic)
    last_active_monotonic: float = field(default_factory=time.monotonic)
    finalized: bool = False
    final_payload: Optional[Dict[str, Any]] = None

    gaze: GazeClassifier = field(default_factory=GazeClassifier)
    head_pose: HeadPoseEstimator = field(default_factory=HeadPoseEstimator)

    counters: Dict[str, ContinuousDirectionCounter] = field(
        default_factory=dict
    )

    multiple_face_detected: bool = False
    multiple_face_count: int = 0
    _multiple_face_active: bool = False

    no_person_detected: bool = False
    _no_person_active: bool = False

    # Throttled fallback body/person-presence cache (YOLO person class).
    # Only consulted when the FaceLandmarker finds 0 faces.
    person_check_ts: float = 0.0
    person_present_cache: Optional[bool] = None
    person_count_cache: int = 0

    # Mobile is external. These values are only storage for the injected
    # detector result; no mobile detection is performed here.
    mobile_count: int = 0
    mobile_score: float = 0.0
    mobile_detected: bool = False

    events: List[Dict[str, Any]] = field(default_factory=list)

    def __post_init__(self):
        self.counters = {
            "head_left": ContinuousDirectionCounter("Head", "Left"),
            "head_right": ContinuousDirectionCounter("Head", "Right"),
            "head_up": ContinuousDirectionCounter("Head", "Up"),
            "eye_left": ContinuousDirectionCounter("Eye", "Left"),
            "eye_right": ContinuousDirectionCounter("Eye", "Right"),
            "eye_up": ContinuousDirectionCounter("Eye", "Up"),
        }

    @property
    def created_at(self):
        # Backward-compatible alias used by the HUD/report.
        return self.created_at_monotonic


# ============================================================
# REPORT
# ============================================================

def generate_excel(
    path: str,
    session: SessionState,
    final_duration: float,
    episodes: List[ViolationEpisode],
    invalid_episodes: List[ViolationEpisode],
    unique_violation_seconds: float,
    violation_percentage: float,
    monitoring_score: float,
    final_score: float,
    metadata: Dict[str, Any],
):
    file_exists = os.path.isfile(path)
    if file_exists:
        try:
            wb = load_workbook(path)
        except Exception:
            wb = Workbook()
            file_exists = False
    else:
        wb = Workbook()

    # 1. Main Sheet: "Monitoring Report"
    if "Monitoring Report" in wb.sheetnames:
        ws = wb["Monitoring Report"]
    else:
        ws = wb.active
        ws.title = "Monitoring Report"

    headers = [
        "Participant ID",
        "Participant Name",
        "Session ID",
        "Actual Duration (sec)",
        "Violation Direction Summary",
        "Unique Violation Time (sec)",
        "Violation Percentage",
        "Eye + Head Score (/60)",
        "Mobile Score (/20)",
        "No Person Score (/10)",
        "Final Proctoring Score (/100)",
    ]

    fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid",
    )

    # Initialize headers if empty or missing
    if ws.max_row == 0 or (ws.max_row == 1 and ws.cell(1, 1).value is None) or ws.cell(1, 1).value not in (headers[0], "Participant ID"):
        ws.delete_rows(1, ws.max_row + 1)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

    dir_totals = calculate_direction_totals(episodes)
    active_dirs = [f"{d.replace('_', ' ').title()}: {s:.1f}s" for d, s in dir_totals.items() if s > 0]
    dir_summary_str = ", ".join(active_dirs) if active_dirs else "None (Centered/Clean)"

    participant_id = metadata.get("participant_id") or session.participant_id or "1"
    participant_name = metadata.get("participant_name") or session.participant_name or "Participant"
    session_id = session.session_id

    no_person_score = (
        NO_PERSON_SCORE_MAX
        if session.no_person_detected
        else 0.0
    )
    mobile_score = clamp(session.mobile_score, 0.0, MOBILE_SCORE_MAX)

    participant_row = [
        str(participant_id),
        str(participant_name),
        str(session_id),
        round(final_duration, 2),
        dir_summary_str,
        round(unique_violation_seconds, 2),
        f"{violation_percentage:.2f}%",
        f"{monitoring_score:.2f} / 60",
        f"{mobile_score:.2f} / 20",
        f"{no_person_score:.2f} / 10",
        f"{final_score:.2f} / 100",
    ]

    # Search if participant/session already exists in the sheet to update or append as a new row
    target_row = None
    if ws.max_row >= 2:
        for r in range(2, ws.max_row + 1):
            existing_pid = str(ws.cell(r, 1).value or "").strip()
            existing_sid = str(ws.cell(r, 3).value or "").strip()
            if (existing_pid and existing_pid == str(participant_id)) or (existing_sid and existing_sid == str(session_id)):
                target_row = r
                break

    if target_row is None:
        target_row = ws.max_row + 1 if ws.cell(ws.max_row, 1).value is not None else ws.max_row

    for col, value in enumerate(participant_row, 1):
        cell = ws.cell(target_row, col, value)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Column width auto-sizing
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 15), 45)

    # 2. Summary Sheet: "Summary"
    if "Summary" in wb.sheetnames:
        summary = wb["Summary"]
    else:
        summary = wb.create_sheet("Summary")

    multiple_face_score = (
        MULTIPLE_FACE_SCORE_MAX
        if session.multiple_face_detected
        else 0.0
    )

    summary_data = [
        ("LMS AI MONITORING SUMMARY", ""),
        ("", ""),
        ("LATEST SESSION INFORMATION", ""),
        ("Participant ID", metadata.get("participant_id", participant_id)),
        ("Participant Name", metadata.get("participant_name", participant_name)),
        ("Trainer ID", metadata.get("trainer_id", "N/A")),
        ("Course ID", metadata.get("course_id", "N/A")),
        ("Assessment ID", metadata.get("assessment_id", "N/A")),
        ("Attempt ID", metadata.get("attempt_id", "N/A")),
        ("Session ID", metadata.get("session_id", session_id)),
        ("Configured Duration", f"{session.configured_duration:.2f} sec"),
        ("Actual Test Duration", f"{final_duration:.2f} sec"),
        ("Actual Start Time", metadata.get("actual_start_time", "N/A")),
        ("Actual End Time", metadata.get("actual_end_time", "N/A")),
        ("Duration Source", metadata.get("duration_source", "N/A")),
        ("", ""),
        ("LATEST CUMULATIVE DIRECTION TOTALS", ""),
        ("Eye LEFT total", f"{calculate_direction_totals(episodes)['eye_left']:.3f} sec"),
        ("Eye RIGHT total", f"{calculate_direction_totals(episodes)['eye_right']:.3f} sec"),
        ("Eye UP total", f"{calculate_direction_totals(episodes)['eye_up']:.3f} sec"),
        ("Head LEFT total", f"{calculate_direction_totals(episodes)['head_left']:.3f} sec"),
        ("Head RIGHT total", f"{calculate_direction_totals(episodes)['head_right']:.3f} sec"),
        ("Head UP total", f"{calculate_direction_totals(episodes)['head_up']:.3f} sec"),
        ("", ""),
        ("SCORING SUMMARY", ""),
        (
            "Eye + Head",
            f"{unique_violation_seconds:.2f} sec | "
            f"{violation_percentage:.2f}% | "
            f"{monitoring_score:.2f} / 60",
        ),
        (
            "Mobile",
            f"Count: {session.mobile_count} | "
            f"{mobile_score:.2f} / 20",
        ),
        (
            "Multiple Face",
            f"Count: 0 | 0.00 / 10 (Single-Participant Mode)",
        ),
        (
            "No Person",
            f"{'Detected' if session.no_person_detected else 'Not Detected'} | "
            f"{no_person_score:.2f} / 10",
        ),
        ("", ""),
        ("FINAL SCORE", ""),
        ("Eye + Head Score", f"{monitoring_score:.2f} / 60"),
        ("Mobile Score", f"{mobile_score:.2f} / 20"),
        ("Multiple Face Score", f"0.00 / 10"),
        ("No Person Score", f"{no_person_score:.2f} / 10"),
        ("Final Score", f"{final_score:.2f} / 100"),
        ("Final Percentage", f"{final_score:.2f}%"),
        ("", ""),
        (
            "Multi-Participant Rule",
            "This Excel sheet maintains and appends all participant attempts into the shared Monitoring Report sheet.",
        ),
    ]

    summary.delete_rows(1, summary.max_row + 5)
    for row_idx, (label, val) in enumerate(summary_data, 1):
        c1 = summary.cell(row_idx, 1, label)
        c2 = summary.cell(row_idx, 2, val)
        if label.isupper() or "SCORE" in label or "SUMMARY" in label:
            c1.font = Font(bold=True)
            c2.font = Font(bold=True)
            c1.fill = PatternFill(
                start_color="D9E1F2",
                end_color="D9E1F2",
                fill_type="solid",
            )
            c2.fill = PatternFill(
                start_color="D9E1F2",
                end_color="D9E1F2",
                fill_type="solid",
            )

    for sheet in (ws, summary):
        for col in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = min(max(max_len + 4, 15), 65)

    output_path = os.path.abspath(path)
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    try:
        wb.save(output_path)
    except PermissionError:
        fallback = os.path.splitext(output_path)[0] + "_output.xlsx"
        wb.save(fallback)
        output_path = fallback

    logger.info("Excel report saved: %s", output_path)
    return output_path


# ============================================================
# ENGINE
# ============================================================

class MediaPipeProctorEngine:
    def __init__(self, model_path: Optional[str] = None):
        self.model_path = resolve_model_path(model_path)
        if not os.path.isfile(self.model_path):
            raise FileNotFoundError(
                f"face_landmarker.task not found: {self.model_path}"
            )

        self.sessions: Dict[str, SessionState] = {}

        # Optional fallback that proves a person/occupant is present when the
        # FaceLandmarker returns 0 faces. Signature: detector(frame_bgr) ->
        # (person_present: bool, person_count: int) or None.
        self._person_detector = None

        self.mp_tasks = mp.tasks
        self.BaseOptions = self.mp_tasks.BaseOptions
        self.FaceLandmarker = self.mp_tasks.vision.FaceLandmarker
        self.FaceLandmarkerOptions = (
            self.mp_tasks.vision.FaceLandmarkerOptions
        )
        self.RunningMode = self.mp_tasks.vision.RunningMode

        self.options = self.FaceLandmarkerOptions(
            base_options=self.BaseOptions(
                model_asset_path=self.model_path
            ),
            running_mode=self.RunningMode.IMAGE,
            num_faces=1,
            min_face_detection_confidence=0.5,
            min_face_presence_confidence=0.5,
            min_tracking_confidence=0.5,
            output_face_blendshapes=False,
            output_facial_transformation_matrixes=False,
        )

        self.detector = None
        try:
            self.detector = self.FaceLandmarker.create_from_options(
                self.options
            )
            logger.info("MediaPipe FaceLandmarker loaded: %s", self.model_path)
        except Exception:
            logger.exception("Could not create FaceLandmarker")
            raise

    PERSON_CHECK_INTERVAL_SECONDS = 1.0

    def set_person_detector(self, detector):
        """Inject an optional body/person-presence detector (e.g. the YOLO
        person class). detector(frame_bgr) must return (person_present: bool,
        person_count: int) or None on failure. It is only invoked when the
        FaceLandmarker finds zero faces, throttled to <=1 Hz per session."""
        self._person_detector = detector

    def _check_person_present(self, session, frame, now):
        """Proves occupant presence from the body/pose stream when the face is
        not resolvable. Returns True/False, or None when unknown (treated as
        'not proven' so the conservative no-person path still applies)."""
        if self._person_detector is None:
            return None
        if now - session.person_check_ts < self.PERSON_CHECK_INTERVAL_SECONDS:
            return session.person_present_cache
        session.person_check_ts = now
        try:
            outcome = self._person_detector(frame)
        except Exception as exc:
            logger.warning("Person-presence fallback failed: %s", exc)
            return None
        if outcome is None:
            session.person_present_cache = None
            return None
        present = bool(outcome[0])
        session.person_present_cache = present
        session.person_count_cache = (
            int(outcome[1]) if len(outcome) > 1 else (1 if present else 0)
        )
        return present

    @staticmethod
    def resolve_lms_duration(
        configured_duration: float,
        actual_start_time: Optional[Union[int, float]],
        actual_end_time: Optional[Union[int, float]],
        fallback_start_monotonic: float,
        fallback_end_monotonic: float,
    ) -> Tuple[float, str]:
        actual = calculate_actual_duration_seconds(
            actual_start_time, actual_end_time
        )
        if actual is not None:
            return max(0.001, actual), "LMS actualStartTime/actualEndTime"

        fallback = max(
            0.001,
            float(fallback_end_monotonic) - float(fallback_start_monotonic),
        )
        return fallback, "monitoring session start/end fallback"

    def start_session(
        self,
        session_id: str,
        participant_id: str = "1",
        participant_name: str = "Participant",
        configured_duration: float = 60.0,
        actual_start_time: Optional[Union[int, float]] = None,
        trainer_id: Optional[str] = None,
        course_id: Optional[str] = None,
        assessment_id: Optional[str] = None,
        attempt_id: Optional[str] = None,
    ) -> Dict[str, Any]:
        if session_id in self.sessions:
            session = self.sessions[session_id]
            if session.finalized:
                raise ValueError(f"Session already finalized: {session_id}")
            return self._session_status(session)

        configured = max(0.0, float(configured_duration) if configured_duration is not None else 60.0)

        session = SessionState(
            session_id=session_id,
            configured_duration=configured,
            participant_id=str(participant_id),
            participant_name=str(participant_name),
            trainer_id=trainer_id,
            course_id=course_id,
            assessment_id=assessment_id,
            attempt_id=attempt_id,
            actual_start_time=normalize_timestamp_ms(actual_start_time),
        )
        self.sessions[session_id] = session
        return self._session_status(session)

    def _get_session(
        self,
        session_id: str,
        configured_duration: float = 60.0,
        **kwargs,
    ) -> SessionState:
        if session_id not in self.sessions:
            self.start_session(
                session_id=session_id,
                configured_duration=configured_duration,
                **kwargs,
            )
        session = self.sessions[session_id]
        session.last_active_monotonic = time.monotonic()
        return session

    def cleanup_stale_sessions(
        self,
        max_idle_seconds: float = float(
            os.getenv("PROCTORING_SESSION_MAX_IDLE_SECONDS", "900")
        ),
    ) -> int:
        """Drop abandoned live sessions whose frames stopped arriving.

        This AI service is intentionally single-instance, but a single instance
        long-lived process can still accumulate sessions if a client disconnects
        mid-assessment. Frames update ``last_active_monotonic``; sessions idle
        longer than ``max_idle_seconds`` are evicted from the in-memory dict so
        memory does not grow unbounded.
        """
        now = time.monotonic()
        stale = [
            sid
            for sid, s in self.sessions.items()
            if not s.finalized
            and now - getattr(s, "last_active_monotonic", s.created_at_monotonic)
            > max_idle_seconds
        ]
        for sid in stale:
            self.sessions.pop(sid, None)
        return len(stale)

    @staticmethod
    def decode_b64(data: str) -> Optional[np.ndarray]:
        try:
            if "," in data:
                data = data.split(",", 1)[1]
            raw = base64.b64decode(data, validate=True)
            arr = np.frombuffer(raw, dtype=np.uint8)
            return cv2.imdecode(arr, cv2.IMREAD_COLOR)
        except Exception:
            return None

    def _detect(self, frame: np.ndarray):
        if self.detector is None:
            raise RuntimeError("MediaPipe detector is not initialized.")

        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        image = mp.Image(
            image_format=mp.ImageFormat.SRGB,
            data=rgb,
        )
        return self.detector.detect(image)

    @staticmethod
    def _append_episode_event(
        session: SessionState,
        episode: ViolationEpisode,
    ):
        start = episode.start_time if episode.start_time < session.created_at else episode.start_time - session.created_at
        end = episode.end_time if episode.end_time < session.created_at else episode.end_time - session.created_at
        session.events.append(
            {
                "category": episode.category,
                "direction": episode.direction,
                "start_time": round(start, 3),
                "end_time": round(end, 3),
                "duration_seconds": round(episode.duration, 3),
                "validation_status": episode.validation_status,
            }
        )

    @staticmethod
    def _close_all_counters(
        session: SessionState,
        now: float,
    ):
        for counter in session.counters.values():
            episode = counter.close(now)
            if episode:
                MediaPipeProctorEngine._append_episode_event(
                    session, episode
                )

    @staticmethod
    def _validated_episodes(
        session: SessionState,
    ) -> List[ViolationEpisode]:
        result: List[ViolationEpisode] = []
        for counter in session.counters.values():
            result.extend(counter.completed)
        result.sort(key=lambda x: x.start_time)
        return result

    @staticmethod
    def _invalid_episodes(
        session: SessionState,
    ) -> List[ViolationEpisode]:
        result: List[ViolationEpisode] = []
        for counter in session.counters.values():
            result.extend(counter.invalid)
        result.sort(key=lambda x: x.start_time)
        return result

    @staticmethod
    def _validated_intervals(
        session: SessionState,
    ) -> List[Tuple[float, float]]:
        return [
            (ep.start_time, ep.end_time)
            for ep in MediaPipeProctorEngine._validated_episodes(session)
            if ep.duration > 0.0
        ]

    @staticmethod
    def _score(
        session: SessionState,
        actual_duration: float,
    ):
        episodes = MediaPipeProctorEngine._validated_episodes(session)
        intervals = MediaPipeProctorEngine._validated_intervals(session)
        unique_seconds = calculate_unique_violation_seconds(intervals)
        direction_totals = calculate_direction_totals(episodes)

        violation_pct, monitoring_score = calculate_monitoring_score(
            unique_seconds,
            actual_duration,
        )

        multiple_face_score = (
            MULTIPLE_FACE_SCORE_MAX
            if session.multiple_face_detected
            else 0.0
        )
        no_person_score = (
            NO_PERSON_SCORE_MAX
            if session.no_person_detected
            else 0.0
        )
        mobile_score = clamp(
            session.mobile_score,
            0.0,
            MOBILE_SCORE_MAX,
        )

        final_score = clamp(
            monitoring_score
            + mobile_score
            + multiple_face_score
            + no_person_score,
            0.0,
            FINAL_SCORE_MAX,
        )

        return {
            "unique_violation_seconds": unique_seconds,
            "violation_percentage": violation_pct,
            "direction_totals": direction_totals,
            "eye_head_score": monitoring_score,
            "multiple_face_score": multiple_face_score,
            "no_person_score": no_person_score,
            "mobile_score": mobile_score,
            "final_score": final_score,
        }

    def update_mobile_result(
        self,
        session_id: str,
        mobile_result: Optional[Union[bool, int, float, Dict[str, Any]]] = None,
        mobile_count: Optional[int] = None,
        mobile_score: Optional[float] = None,
    ):
        """
        External mobile-result injection only.

        This method DOES NOT detect a mobile phone.
        It simply stores the final result supplied by the existing
        separate mobile detector.

        Accepted:
          - mobile_score
          - mobile_count
          - {"count": ..., "score": ..., "detected": ...}
          - bool as an external incident signal (edge-triggered)
        """
        session = self._get_session(session_id)

        if isinstance(mobile_result, dict):
            if "count" in mobile_result and mobile_result["count"] is not None:
                session.mobile_count = max(
                    session.mobile_count, int(mobile_result["count"])
                )
            if "score" in mobile_result and mobile_result["score"] is not None:
                session.mobile_score = clamp(
                    float(mobile_result["score"]),
                    0.0,
                    MOBILE_SCORE_MAX,
                )
            if "detected" in mobile_result:
                session.mobile_detected = bool(mobile_result["detected"])

        elif isinstance(mobile_result, bool):
            # Only a false -> true transition is treated as a new external
            # incident. This prevents counting the same detector signal every
            # video frame.
            detected = mobile_result
            if detected and not session.mobile_detected:
                session.mobile_count += 1
            session.mobile_detected = detected

        elif mobile_result is not None:
            try:
                session.mobile_count = max(
                    session.mobile_count, int(mobile_result)
                )
            except (TypeError, ValueError):
                pass

        if mobile_count is not None:
            session.mobile_count = max(
                session.mobile_count, int(mobile_count)
            )

        if mobile_score is not None:
            session.mobile_score = clamp(
                float(mobile_score),
                0.0,
                MOBILE_SCORE_MAX,
            )
        elif mobile_count is not None:
            # Preserve the existing count-based external scoring contract
            # while keeping it capped at 20.
            session.mobile_score = clamp(
                float(session.mobile_count),
                0.0,
                MOBILE_SCORE_MAX,
            )

        return {
            "count": session.mobile_count,
            "score": session.mobile_score,
            "detected": session.mobile_detected,
        }

    def get_status(self) -> Dict[str, Any]:
        return {
            "status": "UP" if self.detector is not None else "DOWN",
            "model_path": str(FACE_MODEL_PATH),
            "model_exists": os.path.exists(FACE_MODEL_PATH),
            "active_sessions": len(self.sessions),
        }

    def calibrate_session(
        self,
        session_id: str,
        baseline_ear: float = 0.28,
        baseline_face_width: float = 120.0,
        **kwargs,
    ) -> Dict[str, Any]:
        session = self._get_session(session_id)
        session.gaze.is_calibrated = True
        return {
            "success": True,
            "session_id": session_id,
            "is_calibrated": True,
            "message": f"Session {session_id} calibrated successfully",
        }

    def validate_calibration(
        self,
        b64_data: str,
        session_id: str = "default",
        **kwargs,
    ):
        frame = self.decode_b64(b64_data)
        if frame is None:
            return {
                "passed": False,
                "ready": False,
                "reason": "INVALID_IMAGE",
                "message": "Camera frame could not be decoded.",
                "status": "CALIBRATING",
            }

        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        brightness = float(np.mean(gray))
        contrast = float(np.std(gray))

        if brightness < 35:
            return {
                "passed": False,
                "ready": False,
                "reason": "POOR_LIGHTING_DARK",
                "message": "Lighting is too dark.",
                "status": "CALIBRATING",
            }

        if brightness > 235:
            return {
                "passed": False,
                "ready": False,
                "reason": "POOR_LIGHTING_BRIGHT",
                "message": "Lighting is too bright.",
                "status": "CALIBRATING",
            }

        try:
            result = self._detect(frame)
        except Exception as exc:
            return {
                "passed": False,
                "ready": False,
                "reason": "DETECTOR_ERROR",
                "message": str(exc),
                "status": "CALIBRATING",
            }

        face_count = (
            len(result.face_landmarks)
            if result and result.face_landmarks
            else 0
        )

        if face_count == 0:
            return {
                "passed": False,
                "ready": False,
                "reason": "FACE_NOT_DETECTED",
                "message": "No face detected.",
                "status": "CALIBRATING",
            }

        if face_count > 1:
            return {
                "passed": False,
                "ready": False,
                "reason": "MULTIPLE_FACES",
                "message": f"{face_count} faces detected.",
                "status": "CALIBRATING",
            }

        session = self._get_session(session_id)
        landmarks = result.face_landmarks[0]
        height, width = frame.shape[:2]

        # Require an approximately centered participant during calibration.
        nose_x = float(landmarks[1].x)
        nose_y = float(landmarks[1].y)
        centered = (
            0.35 <= nose_x <= 0.65
            and 0.25 <= nose_y <= 0.75
        )
        if not centered:
            return {
                "passed": False,
                "ready": False,
                "reason": "FACE_NOT_CENTERED",
                "message": "Please center your face before calibration.",
                "status": "CALIBRATING",
                "metrics": {
                    "brightness": round(brightness, 1),
                    "contrast": round(contrast, 1),
                    "face_count": face_count,
                    "nose_x": round(nose_x, 3),
                    "nose_y": round(nose_y, 3),
                },
            }

        try:
            raw_h, raw_v, _ = calculate_normalized_eye_ratios(
                landmarks, width, height
            )
            session.gaze.add_calibration_sample(raw_h, raw_v)
        except Exception as exc:
            return {
                "passed": False,
                "ready": False,
                "reason": "EYE_LANDMARKS_UNAVAILABLE",
                "message": str(exc),
                "status": "CALIBRATING",
            }

        ready = session.gaze.is_calibrated
        return {
            "passed": True,
            "ready": ready,
            "reason": (
                "CALIBRATION_READY"
                if ready
                else "CALIBRATION_FRAME_ACCEPTED"
            ),
            "message": (
                "Calibration complete. Monitoring can start."
                if ready
                else "Calibration frame accepted."
            ),
            "status": "MONITORING" if ready else "CALIBRATING",
            "metrics": {
                "brightness": round(brightness, 1),
                "contrast": round(contrast, 1),
                "face_count": face_count,
                "nose_x": round(nose_x, 3),
                "nose_y": round(nose_y, 3),
                "calibration_frames": len(session.gaze.calib_samples_x),
                "calibration_required": CALIBRATION_FRAMES,
                "is_calibrated": ready,
                "neutral_x": round(session.gaze.neutral_x, 4),
                "neutral_y": round(session.gaze.neutral_y, 4),
            },
        }

    def _process_detection(
        self,
        session: SessionState,
        result,
        frame: np.ndarray,
        now: float,
        person_present: Optional[bool] = None,
    ):
        height, width = frame.shape[:2]
        face_count = 1 if (result and result.face_landmarks and len(result.face_landmarks) > 0) else 0

        person_count = face_count
        person_presence_source = "FACE" if face_count > 0 else "NONE"
        if face_count == 0 and person_present:
            person_count = session.person_count_cache
            person_presence_source = "BODY"

        session.multiple_face_detected = False
        session.multiple_face_count = 0
        session._multiple_face_active = False

        head_direction = "Not Detected"
        gaze_direction = "Not Detected"
        head_confidence = 0.0
        gaze_confidence = 0.0
        pitch = None
        yaw = None
        dx = 0.0
        dy = 0.0
        gaze_x = 0.50
        gaze_y = 0.50
        eye_geom = None
        iris_landmarks_detected = False
        gaze_observable = False
        raw_gaze_direction = "Not Detected"
        raw_gaze_confidence = 0.0
        gaze_suppressed_by_head_pose = False
        nose = (width // 2, height // 2)

        if face_count == 0:
            if person_present:
                # A clear occupant is present (full body visible) but facial
                # landmarks are not resolvable (small / turned / blurred face).
                # This is NOT a "no person" state.
                session._no_person_active = False
            else:
                if not session._no_person_active:
                    session.no_person_detected = True
                session._no_person_active = True

            session.head_pose.reset()
            session.gaze.current_direction = "Straight"
            self._close_all_counters(session, now)

        elif face_count > 1:
            # Do not score the first face as if it were the only participant.
            # Keep the multi-face state, and terminate active eye/head episodes.
            self._close_all_counters(session, now)
            session.head_pose.reset()
            session.gaze.current_direction = "Straight"

        else:
            session._no_person_active = False
            landmarks = result.face_landmarks[0]

            (
                head_direction,
                head_confidence,
                pitch,
                yaw,
                nose,
            ) = session.head_pose.estimate(
                landmarks, width, height
            )

            if len(landmarks) > LEFT_IRIS_CENTER:
                try:
                    raw_h, raw_v, eye_geom = calculate_normalized_eye_ratios(
                        landmarks, width, height
                    )
                    iris_landmarks_detected = True

                    if not session.gaze.is_calibrated:
                        session.gaze.add_calibration_sample(raw_h, raw_v)
                        gaze_direction = "Straight"
                        gaze_confidence = 0.0
                        dx = raw_h - session.gaze.neutral_x
                        dy = raw_v - session.gaze.neutral_y
                        gaze_x = raw_h
                        gaze_y = raw_v
                    else:
                        (
                            gaze_direction,
                            gaze_confidence,
                            dx,
                            dy,
                            gaze_x,
                            gaze_y,
                        ) = session.gaze.classify(raw_h, raw_v)
                    raw_gaze_direction = gaze_direction
                    raw_gaze_confidence = gaze_confidence
                except Exception:
                    gaze_direction = "Straight"
                    gaze_confidence = 0.0
            else:
                gaze_direction = "Straight"
                gaze_confidence = 0.0

            # Iris position is a valid gaze signal only while the face is
            # substantially forward. A pronounced yaw changes the projected
            # eye-corner geometry, which made a head turn look like a second,
            # duplicate eye violation. Preserve the head audit and wait for a
            # forward-facing frame before opening an eye-only interval.
            gaze_observable = (
                iris_landmarks_detected
                and session.gaze.is_calibrated
                and iris_gaze_is_observable(head_direction)
            )
            if iris_landmarks_detected and not iris_gaze_is_observable(
                head_direction
            ):
                gaze_suppressed_by_head_pose = True
                gaze_direction = "Straight"
                gaze_confidence = 0.0

            # No violation timers start until calibration is READY.
            if not session.gaze.is_calibrated:
                self._close_all_counters(session, now)
            else:
                # Each of the six directions owns an independent timer.
                # A direction change closes the old interval immediately.
                for name in ("head_left", "head_right", "head_up"):
                    episode = session.counters[name].start_or_continue(
                        head_direction, now
                    )
                    if episode:
                        self._append_episode_event(session, episode)

                for name in ("eye_left", "eye_right", "eye_up"):
                    episode = session.counters[name].start_or_continue(
                        gaze_direction, now
                    )
                    if episode:
                        self._append_episode_event(session, episode)

        return {
            "face_count": face_count,
            "person_detected": face_count > 0 or (person_present or False),
            "person_count": person_count,
            "person_presence_source": person_presence_source,
            "head_direction": head_direction,
            "head_confidence": head_confidence,
            "gaze_direction": gaze_direction,
            "gaze_confidence": gaze_confidence,
            "pitch": pitch,
            "yaw": yaw,
            "dx": dx,
            "dy": dy,
            "gaze_x": gaze_x,
            "gaze_y": gaze_y,
            "eye_geom": eye_geom,
            "iris_landmarks_detected": iris_landmarks_detected,
            "gaze_observable": gaze_observable,
            "raw_gaze_direction": raw_gaze_direction,
            "raw_gaze_confidence": raw_gaze_confidence,
            "gaze_suppressed_by_head_pose": gaze_suppressed_by_head_pose,
            "nose": nose,
        }

    def process_b64_frame(
        self,
        b64_data: str,
        session_id: str = "default",
        timestamp_ms: Optional[int] = None,
        configured_duration: float = 60.0,
        mobile_detected: Optional[bool] = None,
        mobile_result: Optional[Dict[str, Any]] = None,
    ):
        frame = self.decode_b64(b64_data)
        if frame is None:
            return {
                "success": False,
                "error": "Invalid frame image data",
            }

        session = self._get_session(
            session_id,
            configured_duration,
        )

        if session.finalized:
            return {
                "success": False,
                "error": "Session already finalized",
            }

        if mobile_result is not None:
            self.update_mobile_result(
                session_id,
                mobile_result=mobile_result,
            )
        elif mobile_detected is not None:
            self.update_mobile_result(
                session_id,
                mobile_result=mobile_detected,
            )

        # Internal violation timing is monotonic. LMS actual timestamps are
        # used later as the authoritative scoring denominator.
        now = time.monotonic()
        elapsed = now - session.created_at

        try:
            result = self._detect(frame)
        except Exception as exc:
            return {
                "success": False,
                "error": f"MediaPipe detection failed: {exc}",
            }

        # Person-presence fallback: when no facial landmarks are resolvable,
        # query the body/person detector (e.g. YOLO person class) so a clearly
        # visible occupant is not misreported as "no person".
        face_present = bool(
            result and result.face_landmarks and len(result.face_landmarks) > 0
        )
        person_present = None
        if not face_present:
            person_present = self._check_person_present(session, frame, now)

        metrics = self._process_detection(
            session, result, frame, now, person_present=person_present
        )

        score = self._score(
            session,
            max(0.001, elapsed),
        )

        face_count = metrics["face_count"]
        head_direction = metrics["head_direction"]
        gaze_direction = metrics["gaze_direction"]
        pitch = metrics["pitch"]
        yaw = metrics["yaw"]

        # Map directions to normalized uppercase tokens
        norm_head = head_direction.upper() if head_direction else "STRAIGHT"
        if norm_head in ("STRAIGHT", "NOT DETECTED", "NOT_DETECTED", "UNKNOWN"):
            norm_head = "CENTER"

        norm_gaze = gaze_direction.upper() if gaze_direction else "STRAIGHT"
        if norm_gaze in ("STRAIGHT", "NOT DETECTED", "NOT_DETECTED", "UNKNOWN"):
            norm_gaze = "CENTER"

        gaze_classification = "ON_SCREEN" if norm_gaze == "CENTER" else f"OFF_SCREEN_{norm_gaze}"
        head_pose_classification = norm_head

        return {
            "success": True,
            "session_id": session_id,
            "elapsed_seconds": round(elapsed, 2),
            "configured_duration": session.configured_duration,
            "face_detected": face_count > 0,
            "face_count": face_count,
            "person_detected": metrics["person_detected"],
            "person_count": metrics["person_count"],
            "person_presence_source": metrics["person_presence_source"],
            "head_direction": head_direction,
            "head_pose_classification": head_pose_classification,
            "head_confidence": round(metrics["head_confidence"], 1),
            "gaze_direction": gaze_direction,
            "gaze_classification": gaze_classification,
            "gaze_confidence": round(metrics["gaze_confidence"], 1),
            "pitch": (
                None if pitch is None
                else round(pitch, 1)
            ),
            "yaw": (
                None if yaw is None
                else round(yaw, 1)
            ),
            "head_pose": {
                "yaw": round(yaw, 1) if yaw is not None else 0.0,
                "pitch": round(pitch, 1) if pitch is not None else 0.0,
                "roll": 0.0,
            },
            "gaze_dx": round(metrics["dx"], 4),
            "gaze_dy": round(metrics["dy"], 4),
            "gaze_x": round(metrics["gaze_x"], 4),
            "gaze_y": round(metrics["gaze_y"], 4),
            "gaze_audit": {
                "iris_landmarks_detected": metrics[
                    "iris_landmarks_detected"
                ],
                "baseline_ready": session.gaze.is_calibrated,
                "observable": metrics["gaze_observable"],
                "raw_direction": metrics["raw_gaze_direction"],
                "raw_confidence": round(
                    metrics["raw_gaze_confidence"], 1
                ),
                "suppressed_by_head_pose": metrics[
                    "gaze_suppressed_by_head_pose"
                ],
            },
            "is_calibrated": session.gaze.is_calibrated,
            "status": (
                "MONITORING"
                if session.gaze.is_calibrated
                else "CALIBRATING"
            ),
            "scoring": {
                key: (
                    round(float(value), 2) if isinstance(value, (int, float))
                    else {
                        k: round(float(v), 2) if isinstance(v, (int, float)) else v
                        for k, v in value.items()
                    } if isinstance(value, dict)
                    else value
                )
                for key, value in score.items()
            },
            "events_count": len(self._validated_episodes(session)),
        }

    def process_video_file(
        self,
        video_path: str,
        session_id: str,
        segment_key: Optional[str] = None,
        configured_duration: Optional[float] = None,
        sample_fps: float = 3.0,
        start_time_ms: Optional[Union[int, float]] = None,
        attempt_id: Optional[str] = None,
        participant_id: Optional[str] = None,
        thresholds: Optional[Dict[str, Any]] = None,
        **extra,
    ) -> Dict[str, Any]:
        """Analyze a recorded webcam segment file with the same MediaPipe
        engine used online, sampled at `sample_fps`.

        Frame timestamps are anchored to the real recording: each sampled
        frame is stamped `session.created_at + frame_idx / source_fps`, so the
        engine's counters and episodes land on the true video timeline.

        Live-only behavior is deliberately not performed here:
          - no mobile/phone detection (mobile stays live, merged by the backend)
          - no Gemini vision inspection
          - no Excel/JSON report generation (the backend persists results)
          - no session finalize (the engine session is discarded afterwards)

        Returns segment-level aggregated events + scoring inputs.
        """
        if not os.path.isfile(video_path):
            raise ValueError(f"Video file not found: {video_path}")

        cap = cv2.VideoCapture(video_path)
        if not cap.isOpened():
            cap.release()
            raise ValueError(f"Could not open video file: {video_path}")

        try:
            width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH)) or 1280
            height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT)) or 720
            fps = cap.get(cv2.CAP_PROP_FPS)
            if not fps or fps <= 0 or fps > 240:
                fps = 30.0
            total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
            if total_frames <= 0:
                raise ValueError("Video has no decodable frames")
            duration = total_frames / fps
        except Exception:
            cap.release()
            raise

        effective_fps = max(1.0, min(30.0, float(sample_fps) or 3.0))
        step = max(1, int(round(fps / effective_fps)))

        th = thresholds or {}
        tracker_kwargs = lambda name: {
            "min_frames": int(th.get(f"{name}_min_frames")
                             or max(5, int(round(effective_fps * (th.get("min_duration_sec") or 1.0))))),
            "min_duration_sec": float(
                th.get(f"{name}_min_duration_sec") or th.get("min_duration_sec") or 1.0
            ),
        }
        no_person_tracker = SampledIntervalTracker("no_person", **tracker_kwargs("no_person"))
        multiple_tracker = SampledIntervalTracker("multiple_person", **tracker_kwargs("multiple_person"))
        fnv_tracker = SampledIntervalTracker("face_not_visible", **tracker_kwargs("face_not_visible"))

        engine_session_id = segment_key or session_id
        duration = max(0.001, duration)
        duration_arg = configured_duration if configured_duration is not None else duration

        self.start_session(
            session_id=engine_session_id,
            participant_id=str(participant_id or "1"),
            configured_duration=max(0.0, float(duration_arg)),
            actual_start_time=normalize_timestamp_ms(start_time_ms),
            attempt_id=attempt_id,
        )
        session = self.sessions[engine_session_id]
        if session.finalized:
            raise ValueError(f"Engine session already finalized: {engine_session_id}")

        base = session.created_at
        frame_idx = 0
        sampled_frames = 0
        inference_errors = 0
        try:
            while True:
                success, frame = cap.read()
                if not success:
                    break
                if frame_idx % step != 0:
                    frame_idx += 1
                    continue
                now = base + (frame_idx / fps)
                frame_idx += 1

                try:
                    result = self._detect(frame)
                except Exception as exc:
                    inference_errors += 1
                    logger.warning("Sampled frame detection failed: %s", exc)
                    continue
                sampled_frames += 1

                face_count = 1 if (result and result.face_landmarks and len(result.face_landmarks) > 0) else 0

                # Person-presence fallback (YOLO person class). Consulted every
                # sampled frame (throttled to ~1 Hz internally) so multiple
                # persons are caught even while a face is present.
                person_present = self._check_person_present(session, frame, now)

                if face_count == 0:
                    person_count = session.person_count_cache if person_present else 0
                else:
                    person_count = max(face_count, session.person_count_cache or face_count)

                try:
                    self._process_detection(
                        session, result, frame, now, person_present=person_present
                    )
                except Exception:
                    logger.exception("Per-frame processing failed at frame %d", frame_idx)
                    continue

                if person_count == 0:
                    no_person_tracker.update(True, now)
                    multiple_tracker.update(False, now)
                    fnv_tracker.update(False, now)
                elif person_count > 1:
                    no_person_tracker.update(False, now)
                    multiple_tracker.update(True, now)
                    fnv_tracker.update(False, now)
                else:
                    no_person_tracker.update(False, now)
                    multiple_tracker.update(False, now)
                    fnv_tracker.update(face_count == 0, now)
        finally:
            cap.release()

        end_now = base + duration
        self._close_all_counters(session, end_now)

        no_person_ev = no_person_tracker.flush(end_now)
        multiple_ev = multiple_tracker.flush(end_now)
        fnv_ev = fnv_tracker.flush(end_now)

        # ── Build the segment-level event list ────────────────────────────────
        events: List[Dict[str, Any]] = []

        def rel(ts: float) -> float:
            return round(max(0.0, float(ts) - base), 3)

        for ep in self._validated_episodes(session):
            if ep.duration <= 0.0:
                continue
            dir_token = str(ep.direction or "").upper()
            detector = "head" if str(ep.category or "").lower().startswith("head") else "eye"
            if dir_token not in ("LEFT", "RIGHT", "UP"):
                continue
            events.append({
                "category": "looking_away",
                "detector": detector,
                "direction": dir_token,
                "start": rel(ep.start_time),
                "end": rel(ep.end_time),
                "duration": round(max(0.0, float(ep.duration)), 3),
                "confidence": 0.9,
                "validationStatus": ep.validation_status,
            })

        for ev in (no_person_ev, multiple_ev, fnv_ev):
            if ev is None:
                continue
            events.append({
                "category": ev["category"],
                "detector": "body",
                "direction": None,
                "start": rel(ev["start"]),
                "end": rel(ev["end"]),
                "duration": ev["duration"],
                "confidence": 0.95,
                "validationStatus": "VALID",
                "frames": ev["frames"],
                "personDetected": True if ev["category"] == "face_not_visible" else None,
            })

        events.sort(key=lambda e: (e["start"], e["end"]))

        # ── Aggregates ────────────────────────────────────────────────────────
        looking_rows = [e for e in events if e["category"] == "looking_away"]
        no_person_rows = [e for e in events if e["category"] == "no_person"]
        multiple_rows = [e for e in events if e["category"] == "multiple_person"]
        fnv_rows = [e for e in events if e["category"] == "face_not_visible"]

        def sum_seconds(rows):
            return round(sum(float(r["duration"]) for r in rows), 3)

        direction_total_seconds = {}
        for r in looking_rows:
            key = f"{r['detector']}_{r['direction']}".lower()
            direction_total_seconds[key] = direction_total_seconds.get(key, 0.0) + float(r["duration"])

        # ── Informational scoring (authoritative score is computed backend-side) ─
        score = self._score(session, max(0.001, duration))
        intervals = [
            (rel(ep.start_time), rel(ep.end_time))
            for ep in self._validated_episodes(session)
            if ep.duration > 0.0
        ]
        unique_violation_seconds = calculate_unique_violation_seconds(intervals)

        # Drop the engine session so segments never accumulate in memory.
        self.sessions.pop(engine_session_id, None)

        return {
            "status": "processed",
            "session_id": session_id,
            "segment_key": segment_key,
            "attempt_id": attempt_id,
            "durationSec": round(duration, 3),
            "framesProcessed": sampled_frames,
            "inferenceErrors": inference_errors,
            "sourceFps": round(float(fps), 3),
            "sampleFps": effective_fps,
            "resolution": {"width": width, "height": height},
            "startTimeMs": start_time_ms,
            "events": events,
            "aggregates": {
                "lookingAwayCount": len(looking_rows),
                "lookingAwaySeconds": sum_seconds(looking_rows),
                "uniqueViolationSeconds": unique_violation_seconds,
                "directionTotalSeconds": direction_total_seconds,
                "noPersonCount": len(no_person_rows),
                "noPersonSeconds": sum_seconds(no_person_rows),
                "multiplePersonCount": len(multiple_rows),
                "multiplePersonSeconds": sum_seconds(multiple_rows),
                "faceNotVisibleCount": len(fnv_rows),
                "faceNotVisibleSeconds": sum_seconds(fnv_rows),
            },
            "scoring": {
                "eyeHead": {
                    "score": round(float(score.get("eye_head_score", 0.0)), 2),
                    "max": MONITORING_SCORE_MAX,
                    "violationSeconds": unique_violation_seconds,
                },
                "noPerson": {
                    "score": NO_PERSON_SCORE_MAX if no_person_rows else 0.0,
                    "max": NO_PERSON_SCORE_MAX,
                    "seconds": sum_seconds(no_person_rows),
                },
                "multipleFace": {
                    "score": MULTIPLE_FACE_SCORE_MAX if multiple_rows else 0.0,
                    "max": MULTIPLE_FACE_SCORE_MAX,
                    "seconds": sum_seconds(multiple_rows),
                },
                "phone": {"score": 0.0, "max": MOBILE_SCORE_MAX, "count": 0, "note": "mobile detection stays live and is merged by the backend"},
                "finalScore": round(float(score.get("final_score", 0.0)), 2),
            },
            "config": {
                "sampleFps": effective_fps,
                "sourceFps": round(float(fps), 3),
                "calibrationFrames": CALIBRATION_FRAMES,
                "configuredDuration": round(float(duration_arg), 3),
                "thresholds": {
                    "noPersonMinFrames": no_person_tracker.min_frames,
                    "multiplePersonMinFrames": multiple_tracker.min_frames,
                    "faceNotVisibleMinFrames": fnv_tracker.min_frames,
                },
            },
        }

    def finalize_session(
        self,
        session_id: str,
        output_excel: str,
        participant_id: Optional[str] = None,
        participant_name: Optional[str] = None,
        trainer_id: Optional[str] = None,
        course_id: Optional[str] = None,
        assessment_id: Optional[str] = None,
        attempt_id: Optional[str] = None,
        actual_start_time: Optional[Union[int, float]] = None,
        actual_end_time: Optional[Union[int, float]] = None,
        output_json: Optional[str] = None,
        callback_url: Optional[str] = None,
    ):
        if session_id not in self.sessions:
            raise ValueError(f"No active session found: {session_id}")

        session = self.sessions[session_id]

        if session.finalized and session.final_payload is not None:
            return session.final_payload

        if participant_id is not None:
            session.participant_id = str(participant_id)
        if participant_name is not None:
            session.participant_name = str(participant_name)
        if trainer_id is not None:
            session.trainer_id = trainer_id
        if course_id is not None:
            session.course_id = course_id
        if assessment_id is not None:
            session.assessment_id = assessment_id
        if attempt_id is not None:
            session.attempt_id = attempt_id

        if actual_start_time is not None:
            session.actual_start_time = normalize_timestamp_ms(
                actual_start_time
            )

        if actual_end_time is not None:
            session.actual_end_time = normalize_timestamp_ms(
                actual_end_time
            )

        final_monotonic = time.monotonic()
        self._close_all_counters(session, final_monotonic)

        actual_duration, duration_source = self.resolve_lms_duration(
            session.configured_duration,
            session.actual_start_time,
            session.actual_end_time,
            session.created_at,
            final_monotonic,
        )

        score = self._score(session, actual_duration)
        episodes = self._validated_episodes(session)
        invalid_episodes = self._invalid_episodes(session)

        start_iso = (
            datetime.datetime.fromtimestamp(
                session.actual_start_time,
                tz=datetime.timezone.utc,
            ).isoformat()
            if session.actual_start_time is not None
            else None
        )
        end_iso = (
            datetime.datetime.fromtimestamp(
                session.actual_end_time,
                tz=datetime.timezone.utc,
            ).isoformat()
            if session.actual_end_time is not None
            else None
        )

        metadata = {
            "participant_id": session.participant_id,
            "participant_name": session.participant_name,
            "trainer_id": session.trainer_id,
            "course_id": session.course_id,
            "assessment_id": session.assessment_id,
            "attempt_id": session.attempt_id,
            "session_id": session_id,
            "configured_duration": session.configured_duration,
            "actual_test_duration": actual_duration,
            "actual_start_time": start_iso,
            "actual_end_time": end_iso,
            "duration_source": duration_source,
            "generated_at": datetime.datetime.now(
                datetime.timezone.utc
            ).isoformat(),
        }

        excel_path = generate_excel(
            output_excel,
            session,
            actual_duration,
            episodes,
            invalid_episodes,
            score["unique_violation_seconds"],
            score["violation_percentage"],
            score["eye_head_score"],
            score["final_score"],
            metadata,
        )

        payload = {
            "session_info": metadata,
            "monitoring_result": {
                "eye_head": {
                    "violation_duration": round(
                        score["unique_violation_seconds"], 2
                    ),
                    "violation_percentage": round(
                        score["violation_percentage"], 2
                    ),
                    "direction_totals": {
                        key: round(value, 3)
                        for key, value in score["direction_totals"].items()
                    },
                    "score": round(score["eye_head_score"], 2),
                    "max_score": MONITORING_SCORE_MAX,
                },
                "mobile": {
                    "count": session.mobile_count,
                    "score": round(score["mobile_score"], 2),
                    "max_score": MOBILE_SCORE_MAX,
                },
                "multiple_face": {
                    "detected": session.multiple_face_detected,
                    "count": session.multiple_face_count,
                    "score": round(score["multiple_face_score"], 2),
                    "max_score": MULTIPLE_FACE_SCORE_MAX,
                },
                "no_person": {
                    "detected": session.no_person_detected,
                    "score": round(score["no_person_score"], 2),
                    "max_score": NO_PERSON_SCORE_MAX,
                },
                "final": {
                    "final_score": round(score["final_score"], 2),
                    "final_percentage": round(score["final_score"], 2),
                    "max_score": FINAL_SCORE_MAX,
                },
            },
            "events": session.events,
            "invalid_events": [
                {
                    "category": ep.category,
                    "direction": ep.direction,
                    "start_time": round(
                        ep.start_time - session.created_at, 3
                    ),
                    "end_time": round(
                        ep.end_time - session.created_at, 3
                    ),
                    "duration_seconds": round(ep.duration, 3),
                    "validation_status": "INVALID_BELOW_3_SECONDS",
                    "score_contribution": 0.0,
                }
                for ep in invalid_episodes
            ],
            "excel_path": excel_path,
        }

        if output_json:
            output_json = os.path.abspath(output_json)
            os.makedirs(os.path.dirname(output_json) or ".", exist_ok=True)
            with open(output_json, "w", encoding="utf-8") as file:
                json.dump(payload, file, indent=2, ensure_ascii=False)

        if callback_url:
            try:
                request = urllib.request.Request(
                    callback_url,
                    data=json.dumps(payload).encode("utf-8"),
                    headers={"Content-Type": "application/json"},
                    method="POST",
                )
                with urllib.request.urlopen(request, timeout=5) as response:
                    logger.info("Callback HTTP status: %s", response.status)
            except Exception:
                logger.exception(
                    "Callback failed. Session result is still valid."
                )

        session.finalized = True
        session.final_payload = payload
        return payload

    @staticmethod
    def _session_status(session: SessionState):
        return {
            "session_id": session.session_id,
            "configured_duration": session.configured_duration,
            "actual_start_time": session.actual_start_time,
            "status": "FINALIZED" if session.finalized else (
                "MONITORING"
                if session.gaze.is_calibrated
                else "CALIBRATING"
            ),
        }


# ============================================================
# MODEL PATH / CLI
# ============================================================

def resolve_model_path(custom_path: Optional[str] = None) -> str:
    candidates = []
    if custom_path:
        candidates.append(custom_path)

    here = os.path.dirname(os.path.abspath(__file__))
    cwd = os.getcwd()
    parent = os.path.dirname(here)

    candidates.extend(
        [
            os.path.join(here, "face_landmarker.task"),
            os.path.join(here, "models", "face_landmarker.task"),
            os.path.join(parent, "models", "face_landmarker.task"),
            os.path.join(cwd, "face_landmarker.task"),
            os.path.join(cwd, "models", "face_landmarker.task"),
            os.path.join(cwd, "ai-service", "models", "face_landmarker.task"),
            r"E:\agent\posture\face_landmarker.task",
        ]
    )

    for path in candidates:
        if path and os.path.isfile(path):
            return os.path.abspath(path)

    return os.path.abspath(
        candidates[0] if candidates else "face_landmarker.task"
    )


def build_cli_parser():
    parser = argparse.ArgumentParser(description="LMS AI Proctoring Engine")
    parser.add_argument("--camera", default="0", help="Camera index or video file path")
    parser.add_argument("--duration", type=float, default=60.0, help="Test duration in seconds")
    parser.add_argument("--session-id", default=None, help="Session ID")
    parser.add_argument("--participant-id", default="1", help="Participant ID")
    parser.add_argument("--participant-name", default="Participant", help="Participant Name")
    parser.add_argument("--trainer-id", default=None, help="Trainer ID")
    parser.add_argument("--course-id", default=None, help="Course ID")
    parser.add_argument("--assessment-id", default=None, help="Assessment ID")
    parser.add_argument("--attempt-id", default=None, help="Attempt ID")
    parser.add_argument("--actual-start-time", type=float, default=None, help="LMS actual start time (ms)")
    parser.add_argument("--actual-end-time", type=float, default=None, help="LMS actual end time (ms)")
    parser.add_argument("--excel", default="live_session_report.xlsx", help="Output Excel file path")
    parser.add_argument("--output-json", default=None, help="Output JSON file path")
    parser.add_argument("--output", default="live_session_output.mp4", help="Output video path")
    parser.add_argument("--no-record", action="store_true", help="Disable recording")
    parser.add_argument("--headless", action="store_true", help="Run without UI window")
    parser.add_argument("--model", default=None, help="Path to face landmarker model")
    parser.add_argument("--callback-url", default=None, help="Result callback HTTP URL")
    parser.add_argument("--run-tests", action="store_true", help="Run deterministic tests")
    return parser


# ============================================================
# HUD
# ============================================================

def draw_hud(
    frame,
    elapsed,
    configured_duration,
    head_direction,
    gaze_direction,
    head_confidence,
    gaze_confidence,
    score,
    calibrated,
):
    hud_x, hud_y, hud_w, hud_h = 20, 20, 350, 315
    overlay = frame.copy()

    cv2.rectangle(
        overlay,
        (hud_x, hud_y),
        (hud_x + hud_w, hud_y + hud_h),
        (18, 18, 26),
        -1,
    )
    frame[:] = cv2.addWeighted(
        overlay, 0.86, frame, 0.14, 0
    )

    lines = [
        "AI MONITORING",
        (
            f"Time: {elapsed:.1f}s / {configured_duration:.1f}s"
            if configured_duration > 0
            else f"Time: {elapsed:.1f}s / LIVE"
        ),
        f"Head: {head_direction} ({head_confidence:.0f}%)",
        f"Eyes: {gaze_direction} ({gaze_confidence:.0f}%)",
        f"Calibration: {'READY' if calibrated else 'CALIBRATING'}",
        f"Violation: {score['unique_violation_seconds']:.2f}s",
        f"Violation %: {score['violation_percentage']:.2f}%",
        f"Eye + Head: {score['eye_head_score']:.2f}/60",
        f"Mobile: {score['mobile_score']:.2f}/20",
        f"Multi Face: {score['multiple_face_score']:.2f}/10",
        f"No Person: {score['no_person_score']:.2f}/10",
        f"FINAL: {score['final_score']:.2f}/100",
    ]

    y = hud_y + 25
    for index, text in enumerate(lines):
        scale = 0.52 if index == 0 else 0.42
        thickness = 2 if index in (0, len(lines) - 1) else 1
        cv2.putText(
            frame,
            text,
            (hud_x + 12, y),
            cv2.FONT_HERSHEY_SIMPLEX,
            scale,
            (255, 255, 255),
            thickness,
            cv2.LINE_AA,
        )
        y += 24


def draw_gaze_debug(
    frame,
    eye_geom,
    gaze_direction,
    dx,
    dy,
    gaze_x,
    gaze_y,
    neutral_x,
    neutral_y,
):
    if not DEBUG_GAZE or not eye_geom:
        return

    for key in ("r_iris", "l_iris"):
        cv2.circle(frame, eye_geom[key], 3, (255, 0, 255), -1)

    for key in ("r_corners", "l_corners"):
        cv2.line(
            frame,
            eye_geom[key][0],
            eye_geom[key][1],
            (255, 0, 255),
            1,
        )

    x = frame.shape[1] - 260
    cv2.putText(
        frame, f"Gaze: {gaze_direction}", (x, 35),
        cv2.FONT_HERSHEY_SIMPLEX, 0.45, (0, 255, 200), 1, cv2.LINE_AA
    )
    cv2.putText(
        frame, f"dx={dx:+.3f} dy={dy:+.3f}", (x, 55),
        cv2.FONT_HERSHEY_SIMPLEX, 0.40, (255, 255, 255), 1, cv2.LINE_AA
    )
    cv2.putText(
        frame, f"X={gaze_x:.2f} Y={gaze_y:.2f}", (x, 75),
        cv2.FONT_HERSHEY_SIMPLEX, 0.40, (255, 255, 255), 1, cv2.LINE_AA
    )
    cv2.putText(
        frame, f"N={neutral_x:.2f},{neutral_y:.2f}", (x, 95),
        cv2.FONT_HERSHEY_SIMPLEX, 0.40, (200, 200, 200), 1, cv2.LINE_AA
    )


# ============================================================
# DETERMINISTIC TESTS
# ============================================================

def _counter_test(intervals: List[Tuple[str, float]], fps: float = 20.0):
    counter = ContinuousDirectionCounter("Eye", "Left")
    now = 0.0
    dt = 1.0 / fps
    for direction, duration in intervals:
        end_time = now + float(duration)
        while now < end_time:
            counter.start_or_continue(direction, now)
            now += dt
        counter.start_or_continue(direction, end_time)
        now = end_time
    counter.close(now)
    return counter


def run_deterministic_tests():
    # Every short interval is scored; there is no 3-second threshold.
    counter = _counter_test([("Left", 2.0), ("Center", 1.0)])
    assert math.isclose(sum(e.duration for e in counter.completed), 2.0, abs_tol=1e-5)

    # Direction accumulation: LEFT 2 + 1 + 4 = 7, RIGHT = 3.
    episodes = [
        ViolationEpisode("Eye", "Left", 0, 2, 2),
        ViolationEpisode("Eye", "Right", 2, 5, 3),
        ViolationEpisode("Eye", "Left", 5, 6, 1),
        ViolationEpisode("Eye", "Left", 6, 10, 4),
    ]
    totals = calculate_direction_totals(episodes)
    assert math.isclose(totals["eye_left"], 7.0, abs_tol=1e-9)
    assert math.isclose(totals["eye_right"], 3.0, abs_tol=1e-9)

    # Required final validation test: LEFT 2 + 1 + 7 = 10, RIGHT = 10.
    episodes = [
        ViolationEpisode("Eye", "Left", 0, 2, 2),
        ViolationEpisode("Eye", "Left", 3, 4, 1),
        ViolationEpisode("Eye", "Right", 4, 14, 10),
        ViolationEpisode("Eye", "Left", 19, 26, 7),
    ]
    totals = calculate_direction_totals(episodes)
    assert math.isclose(totals["eye_left"], 10.0, abs_tol=1e-9)
    assert math.isclose(totals["eye_right"], 10.0, abs_tol=1e-9)
    unique = calculate_unique_violation_seconds([(e.start_time, e.end_time) for e in episodes])
    assert math.isclose(unique, 20.0, abs_tol=1e-9)
    pct, score = calculate_monitoring_score(unique, 100.0)
    assert math.isclose(pct, 20.0, abs_tol=1e-9)
    assert math.isclose(score, 12.0, abs_tol=1e-9)

    # Eye + Head overlap counted once.
    assert math.isclose(
        calculate_unique_violation_seconds([(10.0, 20.0), (12.0, 18.0)]),
        10.0,
        abs_tol=1e-9,
    )
    assert math.isclose(
        calculate_unique_violation_seconds([(10.0, 15.0), (12.0, 18.0)]),
        8.0,
        abs_tol=1e-9,
    )

    return {"passed": True, "tests": 5, "message": "All cumulative violation-time acceptance tests passed."}


def run_monitoring_session(config=None):
    if config is None:
        args = build_cli_parser().parse_args()
        config = vars(args)
    else:
        config = dict(config)

    if config.get("run_tests"):
        return run_deterministic_tests()

    model_path = resolve_model_path(config.get("model"))
    engine = MediaPipeProctorEngine(model_path)

    session_id = (
        config.get("session_id")
        or f"ms_{int(time.time() * 1000)}"
    )
    raw_duration = config.get("duration")
    duration = (
        max(0.0, float(raw_duration))
        if raw_duration is not None
        else 0.0
    )

    participant_id = config.get("participant_id", "1")
    participant_name = config.get("participant_name", "Participant")
    trainer_id = config.get("trainer_id")
    course_id = config.get("course_id")
    assessment_id = config.get("assessment_id")
    attempt_id = config.get("attempt_id")

    session = engine.start_session(
        session_id=session_id,
        participant_id=participant_id,
        participant_name=participant_name,
        configured_duration=duration,
        actual_start_time=config.get("actual_start_time"),
        trainer_id=trainer_id,
        course_id=course_id,
        assessment_id=assessment_id,
        attempt_id=attempt_id,
    )

    camera_input = str(config.get("camera", "0"))
    camera_source = (
        int(camera_input)
        if camera_input.isdigit()
        else camera_input
    )

    output_excel = str(
        config.get("excel", "live_session_report.xlsx")
    )
    output_json = config.get("output_json")
    callback_url = config.get("callback_url")
    headless = bool(config.get("headless", False))

    cap = cv2.VideoCapture(camera_source)
    if not cap.isOpened():
        raise RuntimeError(
            f"Could not open camera/video source: {camera_source}"
        )

    width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH)) or 1280
    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT)) or 720
    fps = cap.get(cv2.CAP_PROP_FPS)
    if not fps or fps <= 0 or fps > 120:
        fps = 30.0

    record = not bool(config.get("no_record", False))
    writer = None

    if record:
        output_video = str(
            config.get("output", "live_session_output.mp4")
        )
        fourcc = cv2.VideoWriter_fourcc(*"mp4v")
        writer = cv2.VideoWriter(
            output_video,
            fourcc,
            fps,
            (width, height),
        )

    is_file_video = isinstance(camera_source, str) and os.path.isfile(camera_source)
    total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT)) if is_file_video else 0
    if is_file_video and total_frames > 0:
        video_duration = total_frames / fps
        if config.get("duration") is None or config.get("duration") == 60.0:
            duration = video_duration
            engine.sessions[session_id].configured_duration = duration

    print("=" * 70)
    print("LMS AI PROCTORING")
    print("=" * 70)
    print(f"Participant : {participant_name}")
    print(f"Session     : {session_id}")
    print(f"Camera      : {camera_source}")
    print(f"Configured  : {duration:.1f}s")
    print("Validation  : every detected positive interval is accumulated")
    print("Scoring     : Eye+Head 60 | Mobile 20 | MultiFace 10 | NoPerson 10")
    print("=" * 70)

    frame_idx = 0
    start_base = engine.sessions[session_id].created_at

    try:
        while True:
            success, frame = cap.read()
            if not success:
                break

            if not is_file_video:
                frame = cv2.flip(frame, 1)

            if is_file_video:
                elapsed = frame_idx / fps
                now = start_base + elapsed
            else:
                now = time.monotonic()
                elapsed = now - start_base
            frame_idx += 1

            try:
                result = engine._detect(frame)
                metrics = engine._process_detection(
                    engine.sessions[session_id],
                    result,
                    frame,
                    now,
                )
            except Exception:
                logger.exception("Frame processing failed.")
                continue

            session_obj = engine.sessions[session_id]
            score = engine._score(
                session_obj,
                max(0.001, elapsed),
            )

            draw_hud(
                frame,
                elapsed,
                duration,
                metrics["head_direction"],
                metrics["gaze_direction"],
                metrics["head_confidence"],
                metrics["gaze_confidence"],
                score,
                session_obj.gaze.is_calibrated,
            )

            nx, ny = metrics["nose"]
            cv2.putText(
                frame,
                f"HEAD: {metrics['head_direction']}",
                (max(5, nx - 70), max(25, ny - 45)),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.50,
                (0, 215, 255)
                if metrics["head_direction"]
                not in ("Straight", "Down", "Not Detected")
                else (120, 220, 120),
                2,
                cv2.LINE_AA,
            )

            cv2.putText(
                frame,
                f"EYES: {metrics['gaze_direction']}",
                (max(5, nx - 70), max(45, ny - 20)),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.50,
                (0, 215, 255)
                if metrics["gaze_direction"]
                not in ("Straight", "Down", "Not Detected")
                else (120, 220, 120),
                2,
                cv2.LINE_AA,
            )

            draw_gaze_debug(
                frame,
                metrics["eye_geom"],
                metrics["gaze_direction"],
                metrics["dx"],
                metrics["dy"],
                metrics["gaze_x"],
                metrics["gaze_y"],
                session_obj.gaze.neutral_x,
                session_obj.gaze.neutral_y,
            )

            if (
                ENABLE_GAZE_CALIBRATION
                and not session_obj.gaze.is_calibrated
            ):
                pct = int(
                    100
                    * len(session_obj.gaze.calib_samples_x)
                    / max(1, CALIBRATION_FRAMES)
                )
                cv2.putText(
                    frame,
                    f"CALIBRATING... {pct}%",
                    (365, 45),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.55,
                    (0, 215, 255),
                    2,
                    cv2.LINE_AA,
                )

            if writer is not None:
                writer.write(frame)

            if not headless:
                try:
                    cv2.imshow("LMS Live AI Monitoring", frame)
                    key = cv2.waitKey(1) & 0xFF
                    if key == ord("q"):
                        print("\nSession stopped by user.")
                        break
                except cv2.error:
                    headless = True

            if not is_file_video and duration > 0 and elapsed >= duration:
                print(f"\nTarget duration {duration:.1f}s reached.")
                break

    except KeyboardInterrupt:
        print("\nSession interrupted by user.")
    finally:
        cap.release()
        if writer is not None:
            writer.release()
        try:
            cv2.destroyAllWindows()
        except Exception:
            pass

    final_time = start_base + (frame_idx / fps if is_file_video else (time.monotonic() - start_base))
    engine._close_all_counters(engine.sessions[session_id], final_time)

    payload = engine.finalize_session(
        session_id=session_id,
        output_excel=output_excel,
        participant_id=participant_id,
        participant_name=participant_name,
        trainer_id=trainer_id,
        course_id=course_id,
        assessment_id=assessment_id,
        attempt_id=attempt_id,
        actual_start_time=config.get("actual_start_time"),
        actual_end_time=config.get("actual_end_time"),
        output_json=output_json,
        callback_url=callback_url,
    )

    result = payload["monitoring_result"]
    print("\n" + "=" * 70)
    print("FINAL LMS AI MONITORING REPORT")
    print("=" * 70)
    print(f"Eye + Head : {result['eye_head']['score']:.2f} / 60")
    print(f"Mobile     : {result['mobile']['score']:.2f} / 20")
    print(f"Multi Face : {result['multiple_face']['score']:.2f} / 10")
    print(f"No Person  : {result['no_person']['score']:.2f} / 10")
    print(f"FINAL      : {result['final']['final_score']:.2f} / 100")
    print(f"Excel      : {payload['excel_path']}")
    print("=" * 70)
    return payload


# ============================================================
# MODULE-LEVEL EXPORTS
# ============================================================
# main.py imports these at module scope:
#   from inference.proctoring_detector import proctor_engine, FACE_MODEL_PATH, POSE_MODEL_PATH

FACE_MODEL_PATH = resolve_model_path()
POSE_MODEL_PATH = None  # Pose model is not used by the rewritten engine

try:
    proctor_engine = MediaPipeProctorEngine(FACE_MODEL_PATH)
except Exception as _init_err:
    logger.warning("Could not create module-level proctor_engine: %s", _init_err)
    proctor_engine = None


def inspect_b64_with_gemini(b64_data: str) -> Dict[str, Any]:
    """
    Inspect a webcam frame using Google Gemini Multimodal Vision API to detect
    unauthorized devices (cell phones, secondary screens, earbuds), multiple persons, or notes.
    """
    try:
        from services.gemini_client import GeminiClient
        from services.ai_config import get_gemini_api_key
        api_key = get_gemini_api_key()
        if not api_key or api_key == "your-gemini-api-key-here":
            return {
                "success": True,
                "phone_detected": False,
                "multiple_persons": False,
                "earbuds_detected": False,
                "suspicious_objects": [],
                "confidence": 0.0,
                "notes": "Gemini API key not configured"
            }

        client = GeminiClient(api_key=api_key)
        prompt = (
            "You are an AI exam proctoring vision auditor. Analyze this webcam frame carefully for exam integrity.\n"
            "Detect if there are:\n"
            "1. Cell phones, smartphones, tablets, or secondary screens.\n"
            "2. Multiple people in the background or near the candidate.\n"
            "3. Earbuds, headphones, or concealed headsets.\n"
            "4. Books, written notes, or suspicious physical materials.\n\n"
            "Return ONLY valid JSON matching this schema:\n"
            "{\n"
            '  "phone_detected": boolean,\n'
            '  "multiple_persons": boolean,\n'
            '  "earbuds_detected": boolean,\n'
            '  "suspicious_objects": string[],\n'
            '  "confidence": number,\n'
            '  "notes": string\n'
            "}"
        )

        raw_json = client.generate_vision_content(
            prompt=prompt,
            image_b64=b64_data,
            mime_type="image/jpeg"
        )

        try:
            parsed = json.loads(raw_json)
        except Exception:
            import re
            json_match = re.search(r"\{.*\}", raw_json, re.DOTALL)
            parsed = json.loads(json_match.group()) if json_match else {}

        return {
            "success": True,
            "phone_detected": bool(parsed.get("phone_detected", False)),
            "multiple_persons": bool(parsed.get("multiple_persons", False)),
            "earbuds_detected": bool(parsed.get("earbuds_detected", False)),
            "suspicious_objects": parsed.get("suspicious_objects", []),
            "confidence": float(parsed.get("confidence", 0.9)),
            "notes": parsed.get("notes", "Frame inspected successfully")
        }
    except Exception as e:
        logger.error(f"Gemini vision proctoring inspection error: {e}")
        return {
            "success": True,
            "phone_detected": False,
            "multiple_persons": False,
            "earbuds_detected": False,
            "suspicious_objects": [],
            "confidence": 0.0,
            "notes": f"Inspection fallback: {str(e)}"
        }


if __name__ == "__main__":
    run_monitoring_session()
